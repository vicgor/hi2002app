"""Export measurement data to CSV, Excel, JSON, Markdown and PDF."""

from __future__ import annotations

import csv
import json
import logging
from collections.abc import Sequence
from pathlib import Path

from hi2002app.models.measurement import Measurement

logger = logging.getLogger(__name__)


class DataExporter:
    """Stateless helper class that exports a list of Measurements.

    All methods are classmethods so no instance is required.
    """

    @classmethod
    def to_csv(cls, measurements: Sequence[Measurement], path: Path) -> None:
        """Export to CSV."""
        if not measurements:
            logger.warning("No measurements to export.")
            return
        fieldnames = list(measurements[0].to_dict().keys())
        with path.open("w", newline="", encoding="utf-8") as fh:
            writer = csv.DictWriter(fh, fieldnames=fieldnames)
            writer.writeheader()
            for m in measurements:
                writer.writerow(m.to_dict())
        logger.info("Exported %d rows to CSV: %s", len(measurements), path)

    @classmethod
    def to_excel(cls, measurements: Sequence[Measurement], path: Path) -> None:
        """Export to Excel .xlsx using openpyxl."""
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill
        except ImportError as exc:  # pragma: no cover
            raise RuntimeError("openpyxl is required for Excel export") from exc

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Measurements"  # type: ignore[union-attr]

        if not measurements:
            wb.save(path)
            return

        headers = list(measurements[0].to_dict().keys())
        header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)

        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=header)  # type: ignore[union-attr]
            cell.fill = header_fill
            cell.font = header_font

        for row_idx, m in enumerate(measurements, start=2):
            for col_idx, value in enumerate(m.to_dict().values(), start=1):
                ws.cell(  # type: ignore[union-attr]
                    row=row_idx,
                    column=col_idx,
                    value=str(value) if isinstance(value, bool) else value,
                )

        wb.save(path)
        logger.info("Exported %d rows to Excel: %s", len(measurements), path)

    @classmethod
    def to_json(cls, measurements: Sequence[Measurement], path: Path) -> None:
        """Export to JSON."""
        data = [m.to_dict() for m in measurements]
        with path.open("w", encoding="utf-8") as fh:
            json.dump(data, fh, indent=2, ensure_ascii=False)
        logger.info("Exported %d rows to JSON: %s", len(measurements), path)

    @classmethod
    def to_markdown(cls, measurements: Sequence[Measurement], path: Path) -> None:
        """Export to Markdown table."""
        lines: list[str] = ["# HI2002 pH Measurement Data\n"]

        if not measurements:
            lines.append("_No data recorded._")
        else:
            headers = list(measurements[0].to_dict().keys())
            lines.append("| " + " | ".join(headers) + " |")
            lines.append("|" + "---|" * len(headers))
            for m in measurements:
                row = [str(v) for v in m.to_dict().values()]
                lines.append("| " + " | ".join(row) + " |")

        path.write_text("\n".join(lines), encoding="utf-8")
        logger.info("Exported %d rows to Markdown: %s", len(measurements), path)

    @classmethod
    def to_pdf(cls, measurements: Sequence[Measurement], path: Path) -> None:
        """Export to PDF using ReportLab."""
        try:
            from reportlab.lib import colors
            from reportlab.lib.pagesizes import A4
            from reportlab.lib.styles import getSampleStyleSheet
            from reportlab.lib.units import cm
            from reportlab.platypus import (
                Paragraph,
                SimpleDocTemplate,
                Spacer,
                Table,
                TableStyle,
            )
        except ImportError as exc:  # pragma: no cover
            raise RuntimeError("reportlab is required for PDF export") from exc

        doc = SimpleDocTemplate(str(path), pagesize=A4)
        styles = getSampleStyleSheet()
        story = []

        story.append(Paragraph("HI2002 pH Measurement Report", styles["Title"]))
        story.append(Spacer(1, 0.5 * cm))

        if not measurements:
            story.append(Paragraph("No data recorded.", styles["Normal"]))
        else:
            headers = list(measurements[0].to_dict().keys())
            table_data = [
                headers,
                *[[str(v) for v in m.to_dict().values()] for m in measurements],
            ]
            col_width = (A4[0] - 4 * cm) / len(headers)
            table = Table(table_data, colWidths=[col_width] * len(headers))
            table.setStyle(
                TableStyle(
                    [
                        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F4E79")),
                        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                        ("FONTSIZE", (0, 0), (-1, -1), 7),
                        (
                            "ROWBACKGROUNDS",
                            (0, 1),
                            (-1, -1),
                            [colors.white, colors.HexColor("#EBF5FB")],
                        ),
                        ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
                        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                    ]
                )
            )
            story.append(table)

        doc.build(story)
        logger.info("Exported %d rows to PDF: %s", len(measurements), path)
